import openpyxl
import os
import requests
import time
import random
from bs4 import BeautifulSoup
import json

# 导入表格，列表记录第4列
workbook = openpyxl.load_workbook('../各个国家id.xlsx')
sheet = workbook.active  # 第一个表
# 遍历第3列的内容
movie_names = []
# 214-384
for row in sheet.iter_rows(min_row=320,max_row=326,min_col=1, max_col=1, values_only=True):
    movie_names.append(row[0])
workbook.close()

def get_cookies(path):
    # 获取cookies
    f_cookies = open(path, 'r')
    cookies = {}
    for line in f_cookies.read().split(';'):  # 将Cookies字符串其转换为字典
        name, value = line.strip().split('=', 1)
        cookies[name] = value
    return cookies

# 创建一个新的工作簿来保存结果
result_workbook = openpyxl.Workbook()
result_sheet = result_workbook.active
# 写入表头
result_sheet.append(['电影名称', '第一个有效ID'])

if __name__ == '__main__':
    for movie_name in movie_names:
        cookies = get_cookies('cookies.txt')  # cookies文件保存的前面所述的cookies
        url = 'https://search.douban.com/movie/subject_search?search_text=' + str(movie_name) + '&cat=1002'
        header = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36 Edg/134.0.0.0',
            'Connection': 'keep-alive'
        }

        resp = requests.get(url, cookies=cookies, headers=header)
        time.sleep(random.uniform(3, 4))  # 每次请求间隔 1 秒
        resp.encoding = 'UTF-8'
        html = resp.text

        soup = BeautifulSoup(html, 'html.parser')
        script_tag = soup.find('script', {'type': 'text/javascript'})

        # 提取 script 标签中的内容
        script_content = script_tag.string

        # 清理 script 内容，只保留 JSON 数据部分
        start_marker = 'window.__DATA__ = '
        end_marker = ';'
        start_index = script_content.find(start_marker) + len(start_marker)
        end_index = script_content.find(end_marker, start_index)

        # 如果 end_index 为 -1，说明没有找到结束标记，可能需要调整逻辑
        if end_index == -1:
            end_index = len(script_content)  # 直接取到字符串末尾

        json_string = script_content[start_index:end_index].strip()
        print(json_string)  # 打印完整的 JSON 字符串

        try:
            json_data = json.loads(json_string)
        except json.JSONDecodeError as e:
            print("JSON 解析失败:", e)
            print("出错的 JSON 字符串片段:", json_string[e.pos - 50:e.pos + 50])  # 打印出错位置附近的字符串
            continue  # 跳过当前电影，继续下一个


        first_valid_id = None
        # 提取第一个值不为 None 的 id 字段
        for item in json_data['items']:
            item_id = item.get('id')
            if item_id is not None:
                first_valid_id = item_id
                break

        if first_valid_id is not None:
            print(movie_name, "提取的第一个值不为 None 的 id 值:", first_valid_id)
        else:
            print(movie_name, "未找到值不为 None 的 id。")

        # 将数据写入结果 Excel 文件
        result_sheet.append([movie_name, first_valid_id])

    # 保存结果工作簿
    result_workbook.save('result_各国家前10_id.xlsx')
    print("数据已成功保存到result_各国家前10_id.xlsx")