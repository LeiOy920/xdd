#coding = utf-8
import os

import requests
import time
import random
from bs4 import BeautifulSoup

# 遍历表格数据
import openpyxl
workbook = openpyxl.load_workbook(r'C:\Users\86139\PycharmProjects\Database_verify\【重要】地图电影_豆瓣id_m-id【已去重】.xlsx')
sheet = workbook.active # 第一个表
# 遍历第2列的内容__moviename
cell_value=[]
for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
    cell_value.append(row[0])
# # 遍历第3列__m_id
# cell_id=[]
# for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
#     cell_id.append(row[0])
workbook.close()



def get_data(html):
    # 获取所需要的页面数据
    soup = BeautifulSoup(html, 'lxml')
    comment_list = soup.select('.comment > p')
    # next_page = soup.select('#paginator > a')[2].get('href')
    if len(soup.select('#paginator > a')) > 2:
        next_page = soup.select('#paginator > a')[2].get('href')
    else:
        next_page = None
    date_nodes = soup.select('.comment-time')
    return comment_list, next_page, date_nodes

def get_cookies(path):
    # 获取cookies
    f_cookies = open(path, 'r')
    cookies ={}
    for line in f_cookies.read().split(';'): # 将Cookies字符串其转换为字典
        name ,value = line.strip().split('=', 1)
        cookies[name] = value
    return cookies


for i, cell in enumerate(cell_value, start=0):
    j=i+727
    file_path = f"../豆瓣短评/comments{j}.txt"
    if os.path.exists(file_path):
        print(f"{file_path} 已存在，跳过爬取。")
        continue
    abss = 'https://movie.douban.com/subject/'+str(cell) + '/comments'
    firstPag_url = 'https://movie.douban.com/subject/'+str(cell) + '/comments?start=20&limit=20&sort=new_score&status=P&percent_type='
    url = 'https://movie.douban.com/subject/'+str(cell) + '/comments?start=0&limit=20&sort=new_score&status=P'
    header = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 SLBrowser/9.0.6.2081 SLBChan/103 SLBVPV/64-bit',
        'Connection': 'keep-alive'
    }
    if __name__ == '__main__':
        cookies = get_cookies('./cookies.txt')  # cookies文件保存的前面所述的cookies
        print(firstPag_url)
        html = requests.get(firstPag_url, cookies=cookies, headers=header).content
        comment_list, next_page, date_nodes = get_data(html)  # 首先从第一个页面处理
        soup = BeautifulSoup(html, 'lxml')
        # 初始化计数器，用于记录当前爬取的页面数
        page_count = 0

        while next_page and page_count < 4:
            if next_page:
                print(abss + next_page)
            else:
                print("没有下一页了")
                break  # 如果没有下一页，退出循环

            html = requests.get(abss + next_page, cookies=cookies, headers=header).content
            comment_list, next_page, date_nodes = get_data(html)

            with open(file_path, 'a', encoding='utf-8') as f:
                for ind in range(len(comment_list)):
                    comment = comment_list[ind]
                    date = date_nodes[ind]
                    comment = comment.get_text().strip().replace("\n", "")
                    date = date.get_text().strip()
                    f.writelines(date + u'\n' + comment + u'\n')

            # 每成功爬取一页，计数器加 1
            page_count += 1
            time.sleep(1 + float(random.randint(1, 100)) / 20)