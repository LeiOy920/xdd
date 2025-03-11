import os
from openpyxl import load_workbook
from snownlp import SnowNLP

# 存入 excel
def append_row_data_to_excel(row_data):
    try:
        # 尝试加载已有的 Excel 文件
        wb = load_workbook('sentiment_analyse.xlsx')
    except FileNotFoundError:
        # 如果文件不存在，创建新的工作簿和工作表
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        # 添加表头
        header = ['文件名', '平均情感得分', '评论数量', '80 - 100分人数', '60 - 80分人数', '40 - 60分人数', '20 - 40分人数', '0 - 20分人数']
        ws.append(header)
    else:
        ws = wb.active

    # 将 row_data 数组追加到工作表的新行
    ws.append(row_data)
    # 保存工作簿
    wb.save('sentiment_analyse.xlsx')

# 遍历 comment1.txt 至 comment250.txt
for i in range(728,1097):
    file_path = f"../豆瓣短评/comments{i}.txt"
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()
            # 提取偶数行内容（索引从 0 开始，所以索引为奇数的行是偶数行）
            even_lines = [line.strip() for i, line in enumerate(lines) if i % 2 != 0]

            if even_lines:
                score_total = 0  # 分数和
                count = 0  # 分数个数
                scores = []  # 分数列表
                level = [0, 0, 0, 0, 0]  # 各分数等级人数
                name = f'{i}'

                # 对每条评论遍历
                for comment in even_lines:
                    if comment:
                        s = SnowNLP(comment)
                        sentiment = s.sentiments * 100
                        score_total = score_total + sentiment
                        count = count + 1
                        scores.append(sentiment)

                        if 80 <= sentiment <= 100:
                            level[0] = level[0] + 1
                        elif 60 <= sentiment < 80:
                            level[1] = level[1] + 1
                        elif 40 <= sentiment < 60:
                            level[2] = level[2] + 1
                        elif 20 <= sentiment < 40:
                            level[3] = level[3] + 1
                        elif 0 <= sentiment < 20:
                            level[4] = level[4] + 1
                        else:
                            print('分数取值异常')

                if count > 0:
                    row_data = [name, score_total / count, count] + level + scores
                    print(row_data)
                    append_row_data_to_excel(row_data)
                else:
                    print("有效评论数量为 0，无法进行情感分析。")
            else:
                print("评论内容为空，无法进行情感分析。")